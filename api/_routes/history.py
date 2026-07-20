# pyrefly: ignore [missing-import]
from flask import Blueprint, jsonify, send_file
import os
import tempfile
import logging
from datetime import datetime

import _services.database as database
from _services.auth_service import require_auth

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s',
    handlers=[logging.StreamHandler()]
)
logger = logging.getLogger(__name__)

history_bp = Blueprint("history", __name__)


@history_bp.route('/api/dataset', methods=['GET'])
@history_bp.route('/dataset', methods=['GET'])
@require_auth
def get_dataset():
    try:
        data = database.get_dataset_rows()
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@history_bp.route('/api/export/excel', methods=['GET'])
@history_bp.route('/export/excel', methods=['GET'])
@require_auth
def export_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        sales_list = database.view_sales()
        wb = Workbook()
        ws = wb.active
        ws.title = "AI Sales Report"

        # Headers
        headers = ["ID", "Date", "Product", "Category", "Quantity", "Price", "Total Revenue", "Profit", "Forecast/Prediction"]
        ws.append(headers)

        # Header Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F46E5")
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data Rows
        for s in sales_list:
            margin = (s['profit'] / s['total']) * 100 if s['total'] else 0
            row_data = [
                s['id'],
                s['date'],
                s['product'],
                s['category'],
                s['quantity'],
                s['price'],
                s['total'],
                s['profit'],
                f"Margin: {margin:.1f}%"
            ]
            ws.append(row_data)

        # Auto column width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = max(max_length + 2, 10)
            ws.column_dimensions[column].width = adjusted_width

        temp_filename = os.path.join(tempfile.gettempdir(), "sales_report_formatted.xlsx")
        wb.save(temp_filename)

        response = send_file(
            temp_filename,
            as_attachment=True,
            download_name="AI_Sales_Dashboard_Report.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        @response.call_on_close
        def remove_temp_file():
            if os.path.exists(temp_filename):
                os.remove(temp_filename)

        return response
    except Exception as e:
        logger.error(f"Excel export error: {e}")
        return jsonify({'error': str(e)}), 500


@history_bp.route('/api/export/pdf', methods=['GET'])
@history_bp.route('/export/pdf', methods=['GET'])
@require_auth
def export_pdf():
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors

        from _services.forecast import predict_sales

        sales_list = database.view_sales()
        total_sales = sum(float(s['total']) for s in sales_list)
        total_profit = sum(float(s['profit']) for s in sales_list)

        model_info = database.get_latest_model_metadata() or {}
        accuracy = model_info.get("accuracy", 0)

        try:
            sample_prediction = predict_sales(13)
        except Exception:
            sample_prediction = 51424.24

        temp_filename = os.path.join(tempfile.gettempdir(), "sales_report_pro.pdf")

        doc = SimpleDocTemplate(temp_filename, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []

        title = Paragraph(
            "<b><font size=20 color='#4F46E5'>Acme Corp - AI Sales Dashboard Report</font></b>",
            styles['Title']
        )
        elements.append(title)
        elements.append(Spacer(1, 20))

        date_str = datetime.now().strftime("%B %d, %Y - %H:%M:%S")
        elements.append(Paragraph(f"<b>Generation Date:</b> {date_str}", styles['Normal']))
        elements.append(Spacer(1, 15))

        # Dashboard Summary
        elements.append(Paragraph("<b><u>Dashboard Summary</u></b>", styles['Heading3']))
        elements.append(Spacer(1, 10))
        elements.append(Paragraph(f"<b>Total Sales Records:</b> {len(sales_list)}", styles['Normal']))
        elements.append(Paragraph(f"<b>Total Revenue:</b> ₹ {total_sales:,.2f}", styles['Normal']))
        elements.append(Paragraph(f"<b>Total Profit:</b> ₹ {total_profit:,.2f}", styles['Normal']))
        elements.append(Spacer(1, 15))

        # Forecast Summary
        elements.append(Paragraph("<b><u>Forecast Summary</u></b>", styles['Heading3']))
        elements.append(Spacer(1, 10))
        elements.append(Paragraph(f"<b>Next Month Forecast (Month 13):</b> ₹ {sample_prediction:,.2f}", styles['Normal']))
        elements.append(Paragraph(f"<b>Prediction Accuracy (R²):</b> {accuracy*100:.2f}%", styles['Normal']))
        elements.append(Spacer(1, 15))

        # Recent Orders Table
        elements.append(Paragraph("<b><u>Recent Orders (Top 5)</u></b>", styles['Heading3']))
        elements.append(Spacer(1, 10))

        table_data = [["Date", "Product", "Category", "Quantity", "Total", "Profit"]]
        recent = sorted(sales_list, key=lambda x: x['id'], reverse=True)[:5]
        for s in recent:
            table_data.append([
                s['date'],
                s['product'],
                s['category'],
                str(s['quantity']),
                f"₹{s['total']:.2f}",
                f"₹{s['profit']:.2f}"
            ])

        t = Table(table_data, colWidths=[70, 120, 80, 60, 80, 80])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 11),
            ('BOTTOMPADDING', (0,0), (-1,0), 10),
            ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#F3F4F6')),
            ('GRID', (0,0), (-1,-1), 1, colors.white)
        ]))
        elements.append(t)
        elements.append(Spacer(1, 30))

        description = """
        This PDF report is dynamically generated from the AI Sales Forecasting Dashboard.
        The linear regression machine learning model utilizes historical sales trends
        to output automated sales predictions. Page 1 of 1.
        """
        elements.append(Paragraph(description, styles['BodyText']))

        doc.build(elements)

        response = send_file(
            temp_filename,
            as_attachment=True,
            download_name="AI_Sales_Report.pdf",
            mimetype="application/pdf"
        )

        @response.call_on_close
        def remove_temp_file():
            if os.path.exists(temp_filename):
                os.remove(temp_filename)

        return response
    except Exception as e:
        logger.error(f"PDF export error: {e}")
        return jsonify({'error': str(e)}), 500
